
from Util import readcsv, SheetName, stopwordinput, preprocess_spacy, wirteDataList, EmptyCheck, DataListListProcess, DataListList2float
import re
from bert_serving.client import BertClient
from bert_dpcnn import SOPredict, LoadDPCNNRFC, TextRank, TitleSimilarity
import math
from Util import DataListList2float, DataListList2int, StrList2FloatList, StrList2IntList, DataListListProcess, wirteDataList
import pandas as pd
import spacy
import joblib
from sklearn.metrics import accuracy_score, recall_score, f1_score, precision_score, precision_recall_fscore_support
from sklearn.feature_extraction import DictVectorizer
from sklearn.feature_extraction.text import CountVectorizer
import pickle
import openpyxl
#bert-serving-start -model_dir uncased_L-12_H-768_A-12 -num_worker=1

def DataInput(datasetName):
    filepath = "BugSum_Data_"+str(datasetName)+".xlsx"
    sheetNameList = SheetName(filepath)

    dataList = []

    io = pd.io.excel.ExcelFile(filepath, engine='openpyxl')

    for sigSheet in sheetNameList:
        #print("sigSheet", sigSheet)
        data = readcsv(filepath, sigSheet, io)

        dataList.append(data.copy())

    return sheetNameList, dataList

def SigDataInput(dataList, i):
    SenList = dataList[i]["Sentence"]
    lemSenList = dataList[i]["LemedSen"]
    senNumberList = dataList[i]["SenNumber"]
    authorList = dataList[i]["CommentAuthor"]
    senInvolvecd = DataListList2int(dataList[i]["ComSenNum"])
    senVecList = DataListList2float(dataList[i]["SenVec"])
    senOPscore = StrList2FloatList(dataList[i]["SenOP"])
    senTRscore = StrList2FloatList(dataList[i]["SenScore"])
    senTPscore = StrList2FloatList(dataList[i]["TopicScore"])
    senRank = StrList2FloatList(dataList[i]["SenRank"])

    tfidfWordList = dataList[i]["TFIDFWord"]
    tfidfScoreList = StrList2FloatList(dataList[i]["TFIDFScore"])
    ignoreList = StrList2IntList(dataList[i]["IgnoredList"])
    goldenList = StrList2IntList(dataList[i]["GoldenSenNumberList"])

    #print("senVecList", senVecList)
    return SenList, lemSenList, senNumberList, authorList, senInvolvecd, senVecList, senOPscore, senTRscore, senTPscore, senRank, tfidfWordList, tfidfScoreList, ignoreList

def BugSumSigDataInput(dataList, i):
    SenList = dataList[i]["Sentence"]
    lemSenList = dataList[i]["LemedSen"]
    senNumberList = dataList[i]["SenNumber"]
    authorList = dataList[i]["CommentAuthor"]
    senInvolvecd = DataListList2int(dataList[i]["ComSenNum"])
    senVecList = DataListList2float(dataList[i]["SenVec"])
    senOPscore = StrList2FloatList(dataList[i]["SenOP"])
    senTRscore = StrList2FloatList(dataList[i]["SenScore"])
    senTPscore = StrList2FloatList(dataList[i]["TopicScore"])
    senBRscore = StrList2FloatList(dataList[i]["Believability"])
    senRank = StrList2FloatList(dataList[i]["SenRank"])
    tfidfWordList = dataList[i]["TFIDFWord"]
    tfidfScoreList = StrList2FloatList(dataList[i]["TFIDFScore"])
    ignoreList = StrList2IntList(dataList[i]["NewIgnoredList"])
    goldenList = StrList2IntList(dataList[i]["GoldenSenNumberList"])
    evaluationList = DataListList2int(dataList[i]["EvaluationList"])
    evaluationTimeList = StrList2IntList(dataList[i]["EvaluationTimeList"])
    buildInfoMark = DataListList2int(dataList[i]["IgnoredList"])
    fscoreList = StrList2FloatList(dataList[i]["FscoreList"])
    return SenList, lemSenList, senNumberList, authorList, senInvolvecd, senVecList, senOPscore, senTRscore, senTPscore, senBRscore, senRank, tfidfWordList, tfidfScoreList, ignoreList, goldenList, evaluationList, evaluationTimeList, buildInfoMark, fscoreList

def BugSumSigDataWithoutTestInput(dataList, i):
    SenList = dataList[i]["Sentence"]
    lemSenList = dataList[i]["LemedSen"]
    senNumberList = dataList[i]["SenNumber"]
    authorList = dataList[i]["CommentAuthor"]
    senInvolvecd = DataListList2int(dataList[i]["ComSenNum"])
    senVecList = DataListList2float(dataList[i]["SenVec"])
    senOPscore = StrList2FloatList(dataList[i]["SenOP"])
    senTRscore = StrList2FloatList(dataList[i]["SenRank"])
    senTPscore = StrList2FloatList(dataList[i]["TopicScore"])
    tfidfWordList = dataList[i]["TFIDFWord"]
    tfidfScoreList = StrList2FloatList(dataList[i]["TFIDFScore"])
    ignoreList = StrList2IntList(dataList[i]["NewIgnoredList"])
    evaluationList = DataListList2int(dataList[i]["EvaluationList"])
    evaluationTimeList = StrList2IntList(dataList[i]["EvaluationTimeList"])
    buildInfoMark = DataListList2int(dataList[i]["IgnoredList"])
    fscoreList = StrList2FloatList(dataList[i]["FscoreList"])
    return SenList, lemSenList, senNumberList, authorList, senInvolvecd, senVecList, senOPscore, senTRscore, senTPscore, tfidfWordList, tfidfScoreList, ignoreList, evaluationList, evaluationTimeList, buildInfoMark, fscoreList


def quoted(Sentence):
    if len(Sentence)>0 and (Sentence[0]==">" or Sentence.find("> ")!=-1):
        #print("quoted", Sentence)
        return True
    return False

def SignCounter(Sentence):
    counter = 0
    puntc = ['?', ',', '.', '!', ' ', '>']
    for i in Sentence:
        if ((i>='a' and i<='z') or (i>='A' and i<='Z') or i in puntc):
            continue
        else:
            counter = counter + 1
    return counter

def TimeStampRemove(Sentence):
    #result = re.findall('[0-9][0-9][0-9][0-9] [0-9]+[:/][0-9]+[:/][0-9]+',Sentence)
    result = re.findall('[0-9]+[:/-][0-9]+[:/-][0-9]+', Sentence)
    #print("Matched Time", result)
    if len(result)>0:
        return 1
    else:
        return 0

def AnnotatedSentenceCut(Sentence):
    result = re.findall('\*\*\*.*\*\*\*', Sentence)
    #print("Matched Annotated", result)
    if len(result) > 0:
        return 1
    else:
        return 0

def DataRemoveBuildingInfo(dataList, threshold):
    sheetNumber = len(dataList)
    for i in range(sheetNumber):
        #print("dataList[i]", dataList[i])
        dataList[i]["ProcessedSentence"], dataList[i]["IgnoredList"] = SenListRemoveBuildingInfo(dataList[i]["Sentence"], threshold)

    return dataList

def SenListRemoveBuildingInfo(sentences, threshold):
    newSen = sentences.copy()
    ignoredList = []
    senNumber = len(sentences)
    for i in range(senNumber):#ignoredList[i]==1 means that this sentence is building info, and should be ignored during the selection
        ignoredList.append(1)
        #ignoredList.append(0) # 내가 쓴 거
    for i in range(senNumber):
        sigSen = str(sentences[i])
        #print("sigSen", sigSen)
        if ((len(sigSen.split(" "))>threshold or SignCounter(sigSen)>10) and not quoted(sigSen)):
            #print("Question Sentence", sigSen)

            if (TimeStampRemove(sigSen)):
                #print("Time Stamp Cuted Sentence", sigSen)
                continue

            if (AnnotatedSentenceCut(sigSen)):
                #print("Annotated Sentence Cuted Sentence", sigSen)
                continue

            pos = sigSen.find(":")
            if (pos!=-1 and 5<=len(sigSen[:pos].split(" "))<=threshold):
                sigSen=sigSen[:pos]
                #print("PASS After cut Sentence", sigSen)
                newSen[i]=sigSen
                ignoredList[i] = 0
                continue
            #print("Delete Sentence:", len(i.split(" ")), i)
            #print("Cuted Sentence", sigSen)
            #print("cut sentence length", len(sigSen.split(" ")))
            continue
        else:
            #print("PASS Sentence", sigSen)
            newSen[i]=sigSen
            ignoredList[i] = 0
    return newSen, ignoredList

nlp = spacy.load("en_core_web_sm")

def DataLem(dataList, stoplist):
    sheetNumber = len(dataList)
    nlp = spacy.load("en_core_web_sm")
    for i in range(sheetNumber):
        # print("dataList[i]", dataList[i])
        #print("SheetNumber{}、{}".format(i, sheetNumber))
        sentences = dataList[i]["ProcessedSentence"].copy()
        ignoredList = dataList[i]["IgnoredList"].copy()

        #print("sentences", sentences)

        imptSenList = []
        senNum = len(sentences)
        for j in range(senNum):
            sigSen = sentences[j]
            newSen = preprocess_spacy(sigSen, stoplist, nlp)
            if (EmptyCheck(newSen)):
                ignoredList[j] = 1
                imptSenList.append("***EMPTY_SENTENCE***")
            else:
                imptSenList.append(newSen)

        dataList[i]["LemedSen"] = imptSenList.copy()
        dataList[i]["IgnoredList"] = ignoredList.copy()
    return dataList


def Bert2Vec(dataList):
     bc = BertClient(check_length=False)
     sheetLen = len(dataList)
     print("BERTISOK")
     imptSenVecList = []
     for i in range(sheetLen):
         #print("Finish:", i/sheetLen, "Percent")
         targetList = list(dataList[i]["LemedSen"])
         sentenceVecList = bc.encode(targetList)
         #print(sentenceVecList.type())
         sentenceVecList = sentenceVecList.tolist()
         #print("sentenceVecList", len(sentenceVecList[0]))
         dataList[i]["SenVec"] = DataListListProcess(sentenceVecList)
         #print("sentenceVecList", len(sentenceVecList), len(sentenceVecList[0]))
     print("BERTISOK")
     return dataList
def DataOPscore(dataList):
    sheetNum = len(dataList)

    dpcnn, rfc = LoadDPCNNRFC()
    for i in range(sheetNum):
        #print("OPscore Processing:", i, "/", sheetNum)
        senVecList = DataListList2float(dataList[i]["SenVec"])
        #print("senVecList", len(senVecList), len(senVecList[1]))
        OPscoreList = SOPredict(senVecList, dpcnn, rfc)
        dataList[i]["SenOP"] = OPscoreList.tolist()
    return dataList


def TextRankscore(dataList):
    sheetNum = len(dataList)

    #dpcnn, rfc = LoadDPCNNRFC()
    for i in range(sheetNum):
        #print("TextRankscore Processing:", i, "/", sheetNum)
        senVecList = DataListList2float(dataList[i]["SenVec"])

        #print("senVecList", len(senVecList), len(senVecList[1]))

        TextScoreList, TextRankList = TextRank(senVecList)
        dataList[i]["SenScore"] = TextScoreList
        dataList[i]["SenRank"] = TextRankList
    return dataList

def Topicscore(dataList):
    sheetNum = len(dataList)

    bc = BertClient(check_length=False)
    #dpcnn, rfc = LoadDPCNNRFC()
    for i in range(sheetNum):
        titleVec = bc.encode(dataList[i]["title"])
        #print("TitleVec: ", titleVec)
        #print("TextRankscore Processing:", i, "/", sheetNum)
        senVecList = DataListList2float(dataList[i]["SenVec"])


        TopicScoreList = TitleSimilarity(senVecList, titleVec)
        dataList[i]["TopicScore"] = TopicScoreList

    return dataList

def TFIDFCounter(dataList):
    sheetNum = len(dataList)

    tfWordDic = {} #in single file
    idfWordDic = {} #in entire database

    tfDicList = []

    for i in range(sheetNum):
        senList = dataList[i]["LemedSen"]
        tfToalNum = 0
        for sigSen in senList:
            wordList = sigSen.split(" ")
            for sigWord in wordList:
                tfToalNum = tfToalNum + 1
                if sigWord not in tfWordDic:
                    tfWordDic[sigWord] = 1
                    if sigWord not in idfWordDic:
                        idfWordDic[sigWord] = 1
                    else:
                        idfWordDic[sigWord] = idfWordDic[sigWord] + 1
                else:
                    tfWordDic[sigWord] = tfWordDic[sigWord] + 1
        for k, v in tfWordDic.items():
            tfWordDic[k] = float(v)/float(tfToalNum)
        #print("tfWordDic", tfWordDic)
        tfDicList.append(tfWordDic.copy())
        tfWordDic.clear()

    for k, v in idfWordDic.items():
        idfWordDic[k] = math.log(sheetNum/(v+1))
    #print("idfWordDic", idfWordDic)

    for i in range(sheetNum):
        tfidfWordList = []
        tfidfScoreList = []
        for k, v in tfDicList[i].items():
            tfidfWordList.append(k)
            tfidfScoreList.append(v*idfWordDic[k])
        dataList[i]["TFIDFWord"] = tfidfWordList.copy()
        dataList[i]["TFIDFScore"] = tfidfScoreList.copy()
        tfidfWordList.clear()
        tfidfScoreList.clear()

    return dataList

def GoldenNumberMatch(goldenSenNumberList, senNumberList):
    #print("senNumberList", senNumberList)
    #print("goldenSenNumberList", goldenSenNumberList)
    xlsGoldenList = []
    answer = []

    goldenNum = len(goldenSenNumberList)
    for i in range(goldenNum):
        xlsGoldenList.append('\''+goldenSenNumberList[i])
        #xlsGoldenList.append(goldenSenNumberList[i])

    senNum = len(senNumberList)
    for i in range(senNum):
        answer.append(0)

    #print("golden list",  xlsGoldenList)

    for i in xlsGoldenList:
        #print("answer", answer)
        #print("i", type(i))
        #print("senNumberList", senNumberList.index(float(i)))
        #print(senNumberList.index(float(i)))
        answer[senNumberList.index(i)] = 1
    #print("answer", answer)
    return answer

def GoldenSet(dataList, datasetName, sheetNameList):
    filepath = datasetName + "GoldenSet.txt"
    f = open(filepath)
    goldenList = f.readlines()
    startSheetName = goldenList[0][:goldenList[0].find(',')]

    if int(startSheetName) >= 1 and int(startSheetName) <= 9:
        startSheetName = '0' + startSheetName

    #print(startSheetName)

    goldenSenNumberList = []
    GoldenList = []
    #print(goldenList)

    for sigLine in goldenList:
        sheetName = sigLine[:sigLine.find(',')]

        if int(sheetName) >= 1 and int(sheetName) <= 9:
            sheetName = '0' + sheetName

       # print("sheetName", sheetName)
        goldenSenNumber = sigLine[sigLine.find(',')+1:-1]
        goldenSenNumber = goldenSenNumber.strip()
        #print("goldenSenNumber", goldenSenNumber)

        if sheetName!=startSheetName:
           # print("sheetNameList", sheetNameList)
           # print("SheetName", SheetName)

            dataIndex = sheetNameList.index(startSheetName)
            print("Processing:", startSheetName)

            GoldenList = GoldenNumberMatch(goldenSenNumberList, dataList[dataIndex]["SenNumber"])
            #print("GoldenList", GoldenList)
            dataList[dataIndex]["GoldenSenNumberList"]=GoldenList.copy()
            #print("goldenSenNumberList", goldenSenNumberList)
            goldenSenNumberList.clear()

            startSheetName = sheetName

        goldenSenNumberList.append(goldenSenNumber)
    #print("Here")
    dataIndex = sheetNameList.index(startSheetName)
    GoldenList = GoldenNumberMatch(goldenSenNumberList, dataList[dataIndex]["SenNumber"])
    dataList[sheetNameList.index(sheetName)]["GoldenSenNumberList"] = GoldenList.copy()
    #print("Number", sheetNameList.index(sheetName))
    #print("goldenSenNumberList", goldenSenNumberList)
    goldenSenNumberList.clear()
    return dataList

def DataList2Str(dataList):
    sheetLen = len(dataList)
    for i in range(sheetLen):
        imptList = dataList[i]["LemedSen"]
        newList = []
        for sigSen in imptList:
            newList.append(str(sigSen))
        dataList[i]["LemedSen"] = newList.copy()
    return dataList

def DataOPscoreReproSteps(dataList):
    print("DataOPscore: ")
    sheetNum = len(dataList)
    #dpcnn, rfc = LoadDPCNNRFC()
    nlp = spacy.load("en_core_web_sm")
    #vectorizer = DictVectorizer()
    loaded_model = joblib.load('random_forest_repro_steps.pkl')
    # Load the feature vectorizer
    vectorizer = joblib.load('random_forest_vectorizer_repro_steps.pkl')
    #dpcnn = LoadDPCNNRFC()
    #with open("random_forest.pkl", "rb") as f:
    #    loaded_model = pickle.load(f)
    sheetNumber = len(dataList)
    new_data_features = []
    numberBR = 0
    averageAcc = 0
    averageRecall = 0
    averageFscore = 0
    for i in range(sheetNum):
        #print("OPscore Processing:", i, "/", sheetNum)
        #senVecList = DataListList2float(dataList[i]["SenVec"])
        #print(senVecList)
        #exit()
        numberBR += 1
        SenList = dataList[i]["Sentence"]
        new_data_features = [extract_features_repro(str(sentence)) for sentence in SenList]

        # Vectorize the features
        X_new = vectorizer.transform(new_data_features)

        # Get the feature names
        feature_names = vectorizer.get_feature_names()

        # Convert the sparse matrix to a pandas dataframe
        X_new_df = pd.DataFrame.sparse.from_spmatrix(X_new, columns=feature_names)

        # Make predictions on the new data
        predictions = loaded_model.predict(X_new_df)
        dataList[i]["ReproductionStep"] = predictions
        
        goldenList = StrList2IntList(dataList[i]["GoldenRepro"])
        #print(predictions.tolist(),"\ngoldenlist",goldenList)
        #accuracy = precision_score(goldenList, predictions.tolist())
        #recall = recall_score(goldenList, predictions.tolist(), average='weighted')
        #fscore = f1_score(goldenList, predictions.tolist(), average='weighted')
        res = precision_recall_fscore_support(goldenList, predictions.tolist(), average='binary',zero_division=1)
        averageAcc = averageAcc + res[0]
        averageRecall = averageRecall + res[1]
        averageFscore = averageFscore + res[2]
        #acc, recall, f_score = AccuracyMeasure(predictions, goldenList)
        
        dataList[i]["Performance_Repro"] = res#.copy()
    
    averageAcc = averageAcc/numberBR
    averageRecall = averageRecall/numberBR
    #averageFscore = 2 * averageAcc * averageRecall / (averageAcc + averageRecall)
    averageFscore = averageFscore/numberBR
    print("Precision", averageAcc)
    print("Recall", averageRecall)
    print("Fscore", averageFscore)
    return dataList


def DataOPscoreEnv(dataList):
    print("DataOPscore: ")
    sheetNum = len(dataList)
    #dpcnn, rfc = LoadDPCNNRFC()
    nlp = spacy.load("en_core_web_sm")
    #vectorizer = DictVectorizer()
    loaded_model = joblib.load('random_forest_7feat.pkl')
    # Load the feature vectorizer
    vectorizer = joblib.load('vectorizer_ext_7feat.pkl')
    #dpcnn = LoadDPCNNRFC()
    #with open("random_forest.pkl", "rb") as f:
    #    loaded_model = pickle.load(f)
    sheetNumber = len(dataList)
    new_data_features = []
    numberBR = 0
    averageAcc = 0
    averageRecall = 0
    averageFscore = 0
    for i in range(sheetNum):
        #print("OPscore Processing:", i, "/", sheetNum)
        #senVecList = DataListList2float(dataList[i]["SenVec"])
        #print(senVecList)
        #exit()
        numberBR += 1
        SenList = dataList[i]["Sentence"]

        new_data_features = [extract_features_env(str(sentence)) for sentence in SenList]

        # Vectorize the features
        X_new = vectorizer.transform(new_data_features)

        # Get the feature names
        feature_names = vectorizer.get_feature_names()

        # Convert the sparse matrix to a pandas dataframe
        X_new_df = pd.DataFrame.sparse.from_spmatrix(X_new, columns=feature_names)

        # Make predictions on the new data
        predictions = loaded_model.predict(X_new_df)

        
        goldenList = StrList2IntList(dataList[i]["GoldenEnv"])
        #print(predictions.tolist(),"\ngoldenlist",goldenList)
        #accuracy = precision_score(goldenList, predictions.tolist())
        #recall = recall_score(goldenList, predictions.tolist(), average='weighted')
        #fscore = f1_score(goldenList, predictions.tolist(), average='weighted')
        res = precision_recall_fscore_support(goldenList, predictions.tolist(), average='binary',zero_division=1)
        averageAcc = averageAcc + res[0]
        averageRecall = averageRecall + res[1]
        averageFscore = averageFscore + res[2]
        #acc, recall, f_score = AccuracyMeasure(predictions, goldenList)
        dataList[i]["Environment"] = predictions
        dataList[i]["Performance"] = res#.copy()
    
    averageAcc = averageAcc/numberBR
    averageRecall = averageRecall/numberBR
    #averageFscore = 2 * averageAcc * averageRecall / (averageAcc + averageRecall)
    averageFscore = averageFscore/numberBR
    print("Precision", averageAcc)
    print("Recall", averageRecall)
    print("Fscore", averageFscore)
    return dataList


def extract_features_repro(sentence):
    # Parse the sentence using spaCy
    features = {}
    doc = nlp(sentence)
    
    for token in doc:
        if token.pos_ == "VERB":
            for child in token.children:
                if child.dep_ == "dobj":
                    features["verb_obj"] = 1
                else:
                    features["verb_obj"] = 0
    # If no match was found, return None
    # Extract syntactic features
    
    for token in doc:
        if token.pos_ == "VERB" and token.dep_ == "ROOT" and token.tag_ == "VB":
            features["imperative"] = 1
        else:
            features["imperative"] = 0
    
    verb_pattern = re.compile(r"VB|VBD|VBG|VBN")
    # Check if the sentence contains a verb that is a root of a clause
    for token in doc:
        if verb_pattern.match(token.tag_) and token.dep_ == "ROOT":
            features["verbform"] = 1
        else:
            features["verbform"] = 0
    
    if re.match(r"^\d+[.)]", sentence):
        features["numerated"] = 1
    else:
        features["numerated"] = 0

    question = sentence.endswith("?")
    if question == True:
        features["question"] = 0
        features["numerated"] = 0
        features["verbform"] = 0
        features["imperative"] = 0
        features["verb_obj"] = 0

    # Check if the sentence starts with "OS version:"
    #features['num_chars'] = sum(len(token) for token in doc if not token.is_space)
    features['num_words'] = len(doc)
    #features['avg_word_length'] = features['num_chars'] / features['num_words'] if features['num_words'] > 0 else 0

    if len(features) != 6:
        features['extra'] = 0
    return features


def extract_features_env(sentence):
    # Parse the sentence using spaCy
    doc = nlp(sentence)
    flag = False
    # Extract syntactic features
    flag_open = '<details>'
    flag_close = '</details>'
    features = {}
    threshold = 5
   
    if flag_open in sentence:
        flag = True
    if flag_close in sentence:
        flag = False
    if ': ' in doc.text and any(char.isdigit() for char in doc.text) and flag==False:
        features['has_colon'] = 1
    else:
        features['has_colon'] = 0
    
    
    features['num_words'] = len(doc)
    
    if doc.text.startswith('OS version' or 'OS'):
        features['is_os_version'] = 1
    else:
        features['is_os_version'] = 0
    # Check if the sentence starts with "OS version:"
    if (doc.text.startswith('VSCode' or 'VS Code version' or 'VSCode version' or "V8") and features['num_words'] <= threshold):
        features['is_vs_version'] = 1
    else:
        features['is_vs_version'] = 0
    keywords = ["version", 'insider', 'os', 'software', 'generic', 'Linux', "Windows", "OS", 'x64', 'VSCode']
    for keyword in keywords:
        if keyword in doc.text and any(char.isdigit() for char in doc.text) and flag == False:
            features["has_keyword"] = 1
        if keyword in doc.text and features['num_words'] > threshold:
            features["has_keyword"] = 0
    if '|' in sentence and flag == True: #or (doc.text.startswith("Steps" or "flash" or "web")):
        features['symb'] = 0
    else:
        features['symb'] = 1
    for ent in doc.ents:
        if ent.label_ in ['CARDINAL', 'PRODUCT']:
            features['has_version_number'] = 1
            break
        if ent.label_ in ['CARDINAL', 'PRODUCT'] and features['num_words'] > threshold:
            features['has_version_number'] = 0
    
    if re.match(r"^\d+[.)]", sentence):
        features["numerated"] = 0
    else:
        features["numerated"] = 1
    
    question = sentence.endswith("?")
    if question == True:
        features["question"] = 0
    else: 
        features["question"] = 1
    
    if "enabled" in sentence or "|" in sentence:
        features['has_version_number'] = 0
        features["numerated"] = 0
        features["has_keyword"] = 0
        features['is_vs_version'] = 0
        features['is_os_version'] = 0
        features['has_colon'] = 0
    if len(sentence) > 45:
        features['has_version_number'] = 0
        features["numerated"] = 0
        features["has_keyword"] = 0
        features['is_vs_version'] = 0
        features['is_os_version'] = 0
        features['has_colon'] = 0
    if '"' in sentence:
        features['has_version_number'] = 0
        features["numerated"] = 0
        features["has_keyword"] = 0
        features['is_vs_version'] = 0
        features['is_os_version'] = 0
        features['has_colon'] = 0
    
    if len(features) != 9:
        features['extra'] = 0
    return features


def writeClassifSentences():
    xlsx_file = 'BugSum_Data_DDS_35-forrun.xlsx'
    output_folder = 'classifiedsentences\\'
    workbook = openpyxl.load_workbook(xlsx_file)
    # Iterate over each sheet/page in the XLSX file
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Get the column indices for the headers "Sentence" and "GoldenSenNumberList"
        sentence_column_index = None
        golden_sen_column_index = None
        for col_idx, cell in enumerate(sheet[1], start=1):  # Assuming headers are in the first row (row index 1)
            if cell.value == "Sentence":
                sentence_column_index = col_idx
            elif cell.value == "ReproductionStep":
                golden_sen_column_index = col_idx
            elif cell.value == "Environment":
                golden_repro_column_index = col_idx
            elif cell.value == "GoldenEnv":
                golden_env_column_index = col_idx
        if sentence_column_index is not None and golden_sen_column_index is not None and golden_repro_column_index is not None:
            # Get the "Sentence" and "GoldenSenNumberList" columns from the current sheet
            sentencesRepro = []
            sentencesEnv = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row (row index 2)
                sentence = row[sentence_column_index - 1]  # Adjust for 0-based index
                golden_sen_number = row[golden_sen_column_index - 1]  # Adjust for 0-based index
                golden_repro = row[golden_repro_column_index - 1] 
                golden_env = row[golden_env_column_index - 1]
                if golden_sen_number == 1: 
                    sentencesRepro.append(sentence)
                if golden_repro == 1: 
                    sentencesEnv.append(sentence)    
        output_file = output_folder + str(sheet_name) + '.txt'       
        # Write each row of the "Sentence" column to a new line in the TXT file
        with open(output_file, 'w', encoding="utf-8") as f:
            for sentence in sentencesEnv:
                f.write(str(sentence) + "\n")
            for sentence in sentencesRepro:
                f.write(str(sentence) + "\n")
    print("Conversion complete. Each page's content has been written to separate TXT files in", output_folder)


def DataPreprocess(datasetName):
    sheetNameList, dataList = DataInput(datasetName)
    print("Start")
    
    dataList = DataOPscoreReproSteps(dataList) # STep3
    wirteDataList(dataList, sheetNameList, datasetName)
    print("Finish OPscore ReproSteps")

    dataList = DataOPscoreEnv(dataList) # STep3
    wirteDataList(dataList, sheetNameList, datasetName)
    print("Finish OPscore Env")
 
    writeClassifSentences()
    
    dataList = DataRemoveBuildingInfo(dataList, threshold=50)
    print("Finish Remove")

    stoplist = stopwordinput()
    dataList = DataLem(dataList, stoplist)
    wirteDataList(dataList, sheetNameList, datasetName)
    print("Finish Lem")

    dataList = DataList2Str(dataList)

    dataList = Bert2Vec(dataList)
    wirteDataList(dataList, sheetNameList, datasetName)
    print("Finish Bert")

    dataList = TextRankscore(dataList)
    wirteDataList(dataList, sheetNameList, datasetName)
    print("Finish TRscore")
    
    dataList = Topicscore(dataList)
    wirteDataList(dataList, sheetNameList, datasetName)
    print("Finish TPscore")

    dataList = TFIDFCounter(dataList)
    #wirteDataList(dataList, sheetNameList, datasetName)
    print("Finish TFIDF")

    dataList = DataOPscore(dataList)
    wirteDataList(dataList, sheetNameList, datasetName)
    print("Finish OPscore")
 

    wirteDataList(dataList, sheetNameList, datasetName)



if __name__ == "__main__":
    # SDS 데이터셋의 경우
    #DataPreprocess("SDS")
    DataPreprocess("DDS_35-forrun")
    #DataPreprocess("EST")
