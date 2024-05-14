#from idk_util import readcsv, SheetName, stopwordinput, preprocess_spacy, wirteDataList, EmptyCheck, DataListListProcess, DataListList2float
import re
from bert_serving.client import BertClient
#from idk_ import *
import pandas as pd
import math
#from idk_util import DataListList2float, DataListList2int, StrList2FloatList, StrList2IntList, DataListListProcess, wirteDataList
import pandas as pd
import spacy
#from idk_believe import *
from sklearn.metrics import accuracy_score, recall_score, f1_score, precision_score, precision_recall_fscore_support
from sklearn.metrics.pairwise import cosine_similarity
def readcsv(filepath, sheetname, io):
    #df = pd.read_excel(filepath, sheetname, keep_default_na=False)
    df = pd.read_excel(io, sheetname, keep_default_na=False)
    '''
    headers_name = list(df.head())[1:]
    data = {}
    for hname in headers_name:
        data[hname] = EmptyListCheck(list(df[hname]))
    #print("data", data)
    '''
    return df

def writecsv(writer, sheet, data):#sheet is in the form of a list, data is in the form of a dictionary
    df = pd.DataFrame({})
    for (k,j) in data.items():
        df = pd.concat([df, pd.DataFrame({k: j})], axis=1)
    df.to_excel(excel_writer=writer, sheet_name=sheet)


def wirteDataList(dataList, sheetNameList, datasetName):
    filePath = "BugSum_Data_"+datasetName+".xlsx"
    sheetNum = len(sheetNameList)
    writer = pd.ExcelWriter(filePath, engine='xlsxwriter')
    for i in range(sheetNum):
        writecsv(writer, sheetNameList[i], dataList[i])
    #writer.book.use_zip64()
    writer.save()
    writer.close()

def SheetName(filepath):
    #xl = pd.ExcelFile(filepath)
    #xl = pd.read_excel(filepath, engine='openpyxl')
    #print(xl)
    lsheet = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31','32','33','34','35']
    return lsheet
    #return (list(xl.sheet_name))
def DataInput(datasetName):
    filepath = "BugSum_Data_"+str(datasetName)+".xlsx"
    sheetNameList = SheetName(filepath)
    dataList = []

    io = pd.io.excel.ExcelFile(filepath, engine='openpyxl')

    for sigSheet in sheetNameList:
        #print("sigSheet", sigSheet)
        data = readcsv(filepath, sigSheet, io)
        data = data[(data['ReproductionStep'] != 1) & (data['Environment'] != 1)]
        #last_sentence_index = data['Sentence'].last_valid_index()
        # Drop rows below the last row in the "Sentence" column
        #data = data.iloc[:last_sentence_index + 1]
        #print(data)
        #exit()
        #new_data = data.copy()
        
        last_row_index = data['Sentence'].last_valid_index()
        #print(last_row_index)
        data = data.iloc[:last_row_index + 1]
        new_data = pd.concat([data.head(5), data.tail(5)])
        #print(new_data)
        #print(data['SenOP'])
        #print(data['SenOP'][:4])
        data['SenOP'] = data['SenOP'].astype(float)
        data['TopicScore'] = data['TopicScore'].astype(float)
        data['JUST'] = data['SenOP'] * data['TopicScore']
        data = data.sort_values(by='JUST', ascending=False)
        
        #new_data = pd.concat([data.head(5), data.tail(5)])
        num_rows_to_keep = int(0.4 * len(data))
        data = data.head(num_rows_to_keep)
        #print("DATA 1: ", data)
        top_indices = data['JUST'].nlargest(num_rows_to_keep).index
        rows_to_keep = list(top_indices) #+ list(range(5)) + list(range(len(data)-5, len(data)))
        #print(rows_to_keep)
        data = (data.head(len(rows_to_keep)))
        #print("DATA 2: ", data)
        frames = [new_data, data]
        result = pd.concat(frames)
        #print("All: ", result)
        data = result.drop_duplicates(subset=['Sentence'] ,inplace = False)
        #data.drop_duplicates()
        #print("no dups: ", data)
        #exit()
        data = data.sort_values(by="Unnamed: 0")
        #print(result, len(result))
        
        #filtered_df = data.iloc[rows_to_keep]
        #filtered_df = data + new_data
        
        '''
        new_data = pd.concat([data.head(5), data.tail(5)])
        num_rows_to_keep = int(0.4 * len(data))
        data = data.head(num_rows_to_keep)
        #top_indices = data['JUST'].nlargest(num_rows_to_keep).index
        #rows_to_keep = list(top_indices) + list(range(5)) + list(range(len(data)-5, len(data)))
        #filtered_df = data.iloc[rows_to_keep]
        '''
        dataList.append(data.copy())
    return sheetNameList, dataList
    

def wirteDataList(dataList, sheetNameList, datasetName):
    filePath = "BugSum_Data_"+datasetName+".xlsx"
    sheetNum = len(sheetNameList)
    writer = pd.ExcelWriter(filePath, engine='xlsxwriter')
    for i in range(sheetNum):
        writecsv(writer, sheetNameList[i], dataList[i])
    #writer.book.use_zip64()
    writer.save()
    writer.close()

def WriteSentences(datasetName):
    xlsx_file = "BugSum_Data_"+datasetName+".xlsx"
    output_folder = 'ranked_sentences\\'
    import openpyxl

    workbook = openpyxl.load_workbook(xlsx_file)

    # Iterate over each sheet/page in the XLSX file
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Get the "Sentence" column from the current sheet
        sentences = []
        for column in sheet.iter_cols(min_col=2, values_only=True):
            if column[0] == "Sentence":
                for cell in column:
                    if cell != None and cell != "Sentence":
                        sentences.append(cell)
        # Generate the output file name
        output_file = output_folder + str(sheet_name) + '.txt'
        
        # Write each row of the "Sentence" column to a new line in the TXT file
        with open(output_file, 'w', encoding="utf-8") as f:
            for sentence in sentences:
                f.write(str(sentence) + "\n")

    print("Each page's content has been written to separate TXT files in", output_folder)

def DataPreprocess(datasetName):
    sheetNameList, dataList = DataInput(datasetName)
    print("Start")
       
    wirteDataList(dataList, sheetNameList, datasetName)
    WriteSentences(datasetName)



DataPreprocess("DDS_35-forrun")











