import pandas as pd

# Read the XLSX file
xlsx_file = 'BugSum_Data_DDS_new.xlsx'
output_folder = 'dds_new_ranked\\'
import openpyxl

workbook = openpyxl.load_workbook(xlsx_file)

# Iterate over each sheet/page in the XLSX file
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    # Get the "Sentence" column from the current sheet
    sentences = []
    for column in sheet.iter_cols(min_col=2, values_only=True):
        if column[0] == "Sentence":
            for cell in column:
                if cell != None and cell != "Sentence":
                    sentences.append(cell)
    # Generate the output file name
    output_file = output_folder + str(sheet_name) + '.txt'
    
    # Write each row of the "Sentence" column to a new line in the TXT file
    with open(output_file, 'w', encoding="utf-8") as f:
        for sentence in sentences:
            f.write(str(sentence) + "\n")

print("Conversion complete. Each page's content has been written to separate TXT files in", output_folder)

